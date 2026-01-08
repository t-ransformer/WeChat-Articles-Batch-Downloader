#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从Excel文件中提取文章URL
用于处理从 wechat-article-exporter 导出的Excel文件
"""
import openpyxl
import re
import sys
import os
import tempfile
from urllib.parse import urlparse


def extract_urls_from_excel(excel_file, output_file='urls.txt'):
    """
    从Excel文件中提取所有文章URL
    
    Args:
        excel_file: Excel文件路径
        output_file: 输出文件路径（默认urls.txt）
    """
    print(f"正在读取Excel文件: {excel_file}")
    
    if not os.path.exists(excel_file):
        print(f"错误: 文件不存在 {excel_file}")
        return
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        print(f"工作表列表: {wb.sheetnames}")
        
        all_urls = []
        
        # 遍历所有工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n处理工作表: {sheet_name}")
            
            # 查找包含URL的列
            # 通常URL在"链接"、"URL"、"文章链接"等列中
            url_column = None
            header_row = None
            
            # 查找表头行（通常在第一行或前几行）
            for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=False), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    if cell.value:
                        cell_value = str(cell.value).lower()
                        # 查找包含"url"、"链接"、"link"等关键词的列
                        if any(keyword in cell_value for keyword in ['url', '链接', 'link', '文章链接', 'article']):
                            url_column = col_idx
                            header_row = row_idx
                            print(f"  找到URL列: 第{col_idx}列 (表头: {cell.value})")
                            break
                if url_column:
                    break
            
            if not url_column:
                print(f"  警告: 未找到URL列，尝试扫描所有单元格...")
                # 如果没有找到明确的URL列，扫描所有单元格
                for row in ws.iter_rows(values_only=True):
                    for cell_value in row:
                        if cell_value:
                            url = extract_url_from_text(str(cell_value))
                            if url:
                                all_urls.append(url)
            else:
                # 从找到的URL列提取
                for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                    if url_column <= len(row):
                        cell_value = row[url_column - 1]
                        if cell_value:
                            url = extract_url_from_text(str(cell_value))
                            if url:
                                all_urls.append(url)
                                print(f"  第{row_idx}行: {url[:50]}...")
        
        # 去重
        unique_urls = list(dict.fromkeys(all_urls))
        
        print(f"\n总共提取到 {len(unique_urls)} 个唯一URL")
        
        # 保存到文件
        with open(output_file, 'w', encoding='utf-8') as f:
            for url in unique_urls:
                f.write(url + '\n')
        
        if output_file != os.path.join(tempfile.gettempdir(), 'wechat_urls_temp.txt'):
            print(f"\nURL列表已保存到: {output_file}")
        
        return unique_urls
        
    except Exception as e:
        print(f"错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


def extract_url_from_text(text):
    """
    从文本中提取微信公众号文章URL
    
    Args:
        text: 文本内容
        
    Returns:
        str: 提取的URL，如果没有则返回None
    """
    # 匹配微信公众号文章URL
    pattern = r'https?://mp\.weixin\.qq\.com/s/[A-Za-z0-9_-]+'
    match = re.search(pattern, text)
    
    if match:
        url = match.group(0)
        # 清理URL，移除可能的参数
        try:
            parsed = urlparse(url)
            clean_url = f"{parsed.scheme}://{parsed.netloc}{parsed.path}"
            return clean_url
        except:
            return url
    
    return None


def main():
    """主函数"""
    if len(sys.argv) < 2:
        print("使用方法:")
        print("  python extract_urls_from_excel.py <Excel文件路径> [输出文件路径]")
        print("\n示例:")
        print("  python extract_urls_from_excel.py 微信公众号文章.xlsx")
        print("  python extract_urls_from_excel.py 微信公众号文章.xlsx urls.txt")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else 'urls.txt'
    
    extract_urls_from_excel(excel_file, output_file)


if __name__ == '__main__':
    main()

